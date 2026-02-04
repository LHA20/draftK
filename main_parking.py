#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Parking Management System - Complete Version with Serial Port Support
Integrated YOLO license plate detection, EasyOCR text extraction, and Arduino/ESP32 RFID reader support
"""

import sys
import os
from datetime import datetime, timedelta
from pathlib import Path
import threading
import csv
import json
import serial
import serial.tools.list_ports

import cv2
import numpy as np
import torch

# Monkey-patch torch.load to disable security check for YOLO model loading
_original_torch_load = torch.load
def patched_torch_load(f, *args, **kwargs):
    kwargs['weights_only'] = False
    return _original_torch_load(f, *args, **kwargs)
torch.load = patched_torch_load

from ultralytics import YOLO
import easyocr

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QHBoxLayout, QVBoxLayout,
    QPushButton, QLabel, QTableWidget, QTableWidgetItem, QLineEdit,
    QMessageBox, QHeaderView, QFileDialog, QComboBox, QGroupBox, QDialog,
    QScrollArea
)
from PyQt5.QtGui import QImage, QPixmap, QFont, QColor
from PyQt5.QtCore import Qt, QTimer, pyqtSignal, QObject

# Pre-load OCR reader to avoid delays during first capture
print("Loading OCR reader...", end=" ", flush=True)
OCR_READER = easyocr.Reader(['en'], gpu=False)
print("✓")


class CameraSignals(QObject):
    """Signal emitter for camera worker thread events"""
    frame_ready = pyqtSignal(np.ndarray)  # Emitted when new frame is captured
    error_signal = pyqtSignal(str)  # Emitted when camera error occurs


class SerialSignals(QObject):
    """Signal emitter for serial port worker thread events"""
    data_received = pyqtSignal(dict)  # Emitted when JSON card scan data received
    connection_status = pyqtSignal(bool, str)  # Emitted when connection status changes (connected, message)
    error_signal = pyqtSignal(str)  # Emitted when serial communication error occurs


class CameraWorker(threading.Thread):
    """Worker thread for continuous 30 FPS camera capture and YOLO inference"""
    
    def __init__(self, signals, models_path):
        super().__init__(daemon=True)
        self.signals = signals
        self.is_running = False
        self.cap = None
        self.license_plate_detector = None
        
        # Load YOLO model for license plate detection
        try:
            self.license_plate_detector = YOLO(str(models_path / 'license_plate_detector.pt'))
            print("✓ License plate detector model loaded")
        except Exception as e:
            self.signals.error_signal.emit(f"Error loading YOLO model: {e}")
    
    def run(self):
        """Continuously capture frames from webcam at 30 FPS"""
        try:
            self.is_running = True
            self.cap = cv2.VideoCapture(0)
            
            if not self.cap.isOpened():
                self.signals.error_signal.emit("Cannot open camera")
                return
            
            # Set video capture parameters
            self.cap.set(cv2.CAP_PROP_FRAME_WIDTH, 1280)
            self.cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 720)
            self.cap.set(cv2.CAP_PROP_FPS, 30)
            
            print("✓ Camera initialized")
            
            while self.is_running:
                ret, frame = self.cap.read()
                if not ret:
                    break
                
                # Mirror the frame for better UX
                frame = cv2.flip(frame, 1)
                self.signals.frame_ready.emit(frame)
        
        except Exception as e:
            self.signals.error_signal.emit(f"Camera error: {str(e)}")
        finally:
            if self.cap:
                self.cap.release()
    
    def stop(self):
        """Stop the camera thread gracefully"""
        self.is_running = False


class SerialWorker(threading.Thread):
    """Worker thread for serial port communication with Arduino/ESP32 RFID reader"""
    
    def __init__(self, signals):
        super().__init__(daemon=True)
        self.signals = signals
        self.is_running = False
        self.serial_port = None
        self.port_name = None
        self.baud_rate = None
    
    def set_connection(self, port_name, baud_rate):
        """Configure port and baud rate before connection"""
        self.port_name = port_name
        self.baud_rate = baud_rate
    
    def run(self):
        """Listen for incoming JSON data from serial port"""
        try:
            if not self.port_name or not self.baud_rate:
                self.signals.error_signal.emit("Port or baud rate not set")
                return
            
            self.serial_port = serial.Serial(
                port=self.port_name,
                baudrate=self.baud_rate,
                timeout=1,
                write_timeout=1
            )
            
            self.is_running = True
            self.signals.connection_status.emit(True, f"Connected to {self.port_name} @ {self.baud_rate}")
            print(f"✓ Serial port connected: {self.port_name} @ {self.baud_rate} baud")
            
            while self.is_running:
                try:
                    if self.serial_port.in_waiting > 0:
                        line = self.serial_port.readline().decode('utf-8').strip()
                        
                        if line:
                            try:
                                # Parse JSON data from Arduino/ESP32
                                data = json.loads(line)
                                # Validate JSON contains required fields
                                if 'event' in data and 'uid' in data:
                                    self.signals.data_received.emit(data)
                                    print(f"Card scanned: {data.get('uid')} - Gate: {data.get('gate')}")
                            except json.JSONDecodeError as e:
                                self.signals.error_signal.emit(f"Invalid JSON: {line}")
                
                except Exception as e:
                    if self.is_running:
                        self.signals.error_signal.emit(f"Serial read error: {str(e)}")
                    break
        
        except serial.SerialException as e:
            self.signals.error_signal.emit(f"Serial connection error: {str(e)}")
            self.signals.connection_status.emit(False, "Disconnected")
        except Exception as e:
            self.signals.error_signal.emit(f"Serial worker error: {str(e)}")
        finally:
            self.is_running = False
            if self.serial_port and self.serial_port.is_open:
                self.serial_port.close()
            self.signals.connection_status.emit(False, "Disconnected")
    
    def stop(self):
        """Stop serial worker and close port gracefully"""
        self.is_running = False
        if self.serial_port and self.serial_port.is_open:
            self.serial_port.close()


def get_available_ports():
    """Auto-detect available serial ports from connected devices"""
    ports = []
    for port_info in serial.tools.list_ports.comports():
        ports.append((port_info.device, port_info.description))
    return ports



def extract_license_plate(crop):
    """Extract license plate text using OCR with preprocessing"""
    try:
        if crop.size == 0:
            return None, None
        
        # Convert BGR to Grayscale
        if len(crop.shape) == 3:
            gray = cv2.cvtColor(crop, cv2.COLOR_BGR2GRAY)
        else:
            gray = crop
        
        # Apply CLAHE (Contrast Limited Adaptive Histogram Equalization)
        clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
        gray = clahe.apply(gray)
        
        # Apply binary threshold with Otsu's method
        _, binary = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
        
        # Run EasyOCR on preprocessed image
        results = OCR_READER.readtext(binary, detail=1)
        
        if not results:
            return None, None
        
        # Combine extracted text and calculate average confidence
        full_text = ""
        total_score = 0.0
        
        for (bbox, text, score) in results:
            clean_text = text.upper().strip().replace(' ', '').replace('.', '')
            clean_text = ''.join(c for c in clean_text if c.isalnum())
            
            if clean_text:
                full_text += clean_text
                total_score += score
        
        # Return only if extracted text has minimum length (license plate)
        if full_text and len(full_text) >= 4:
            avg_score = total_score / len(results) if results else 0.0
            return full_text.upper(), avg_score
        
        return None, None
    
    except Exception as e:
        print(f"OCR Error: {e}")
        return None, None


class SettingsDialog(QDialog):
    """Settings Dialog for Serial Port Configuration - Material Design"""
    
    def __init__(self, parent, current_port=None, current_baud='9600'):
        super().__init__(parent)
        self.setWindowTitle("Serial Port Settings")
        self.setGeometry(200, 200, 500, 300)
        self.setStyleSheet(self.get_material_style())
        
        self.current_port = current_port
        self.current_baud = current_baud
        
        self.init_ui()
    
    def get_material_style(self):
        """Material Design stylesheet"""
        return """
            QDialog {
                background-color: #fafafa;
            }
            QLabel {
                color: #212121;
                font-weight: 500;
            }
            QComboBox {
                border: 1px solid #bdbdbd;
                border-radius: 4px;
                padding: 8px;
                background-color: #ffffff;
                color: #212121;
                selection-background-color: #2196F3;
            }
            QComboBox:focus {
                border: 2px solid #2196F3;
                padding: 7px;
            }
            QPushButton {
                background-color: #2196F3;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 10px 24px;
                font-weight: bold;
                min-height: 36px;
            }
            QPushButton:hover {
                background-color: #1976D2;
            }
            QPushButton:pressed {
                background-color: #1565C0;
            }
            QPushButton#CancelButton {
                background-color: #757575;
            }
            QPushButton#CancelButton:hover {
                background-color: #616161;
            }
        """
    
    def init_ui(self):
        """Initialize settings dialog UI"""
        layout = QVBoxLayout()
        layout.setSpacing(20)
        layout.setContentsMargins(24, 24, 24, 24)
        
        # Title
        title = QLabel("Serial Port Configuration")
        title.setFont(QFont("Arial", 14, QFont.Bold))
        title.setStyleSheet("color: #1976D2;")
        layout.addWidget(title)
        
        # Port selection
        port_layout = QHBoxLayout()
        port_label = QLabel("Serial Port:")
        port_label.setMinimumWidth(120)
        port_label.setFont(QFont("Arial", 10, QFont.Bold))
        port_layout.addWidget(port_label)
        
        self.port_combo = QComboBox()
        self.port_combo.setMinimumHeight(36)
        ports = get_available_ports()
        if ports:
            for port_name, port_desc in ports:
                self.port_combo.addItem(f"{port_name} - {port_desc}", port_name)
        else:
            self.port_combo.addItem("No ports found", None)
        
        if self.current_port:
            idx = self.port_combo.findData(self.current_port)
            if idx >= 0:
                self.port_combo.setCurrentIndex(idx)
        
        port_layout.addWidget(self.port_combo)
        layout.addLayout(port_layout)
        
        # Baud rate selection
        baud_layout = QHBoxLayout()
        baud_label = QLabel("Baud Rate:")
        baud_label.setMinimumWidth(120)
        baud_label.setFont(QFont("Arial", 10, QFont.Bold))
        baud_layout.addWidget(baud_label)
        
        self.baud_combo = QComboBox()
        self.baud_combo.setMinimumHeight(36)
        baud_rates = ['9600', '19200', '38400', '57600', '115200']
        self.baud_combo.addItems(baud_rates)
        if self.current_baud in baud_rates:
            self.baud_combo.setCurrentText(self.current_baud)
        
        baud_layout.addWidget(self.baud_combo)
        layout.addLayout(baud_layout)
        
        layout.addStretch()
        
        # Buttons
        button_layout = QHBoxLayout()
        button_layout.setSpacing(12)
        
        cancel_btn = QPushButton("Cancel")
        cancel_btn.setObjectName("CancelButton")
        cancel_btn.setMinimumHeight(36)
        cancel_btn.setMaximumWidth(100)
        cancel_btn.clicked.connect(self.reject)
        button_layout.addStretch()
        button_layout.addWidget(cancel_btn)
        
        ok_btn = QPushButton("OK")
        ok_btn.setMinimumHeight(36)
        ok_btn.setMaximumWidth(100)
        ok_btn.clicked.connect(self.accept)
        button_layout.addWidget(ok_btn)
        
        layout.addLayout(button_layout)
        
        self.setLayout(layout)
    
    def get_settings(self):
        """Return selected settings"""
        return {
            'port': self.port_combo.currentData(),
            'baud': int(self.baud_combo.currentText())
        }


class CheckInConfirmationDialog(QDialog):
    """Large, professional check-in confirmation dialog with better UI"""
    
    def __init__(self, parent, card_id, license_plate, slot, time_str):
        super().__init__(parent)
        self.setWindowTitle("🚗 CHECK IN CONFIRMATION")
        self.setGeometry(200, 200, 600, 400)
        self.setStyleSheet("background-color: #f5f5f5;")
        self.user_confirmed = False
        
        self.init_ui(card_id, license_plate, slot, time_str)
    
    def init_ui(self, card_id, license_plate, slot, time_str):
        """Initialize confirmation dialog UI with large fonts"""
        layout = QVBoxLayout()
        layout.setSpacing(20)
        layout.setContentsMargins(24, 24, 24, 24)
        
        # Title with icon
        title_layout = QHBoxLayout()
        title = QLabel("✓ LICENSE PLATE DETECTED")
        title.setFont(QFont("Arial", 16, QFont.Bold))
        title.setStyleSheet("color: #2e7d32;")
        title_layout.addWidget(title)
        title_layout.addStretch()
        layout.addLayout(title_layout)
        
        # Separator line
        separator = QLabel()
        separator.setStyleSheet("background-color: #90caf9; height: 2px;")
        separator.setMinimumHeight(2)
        layout.addWidget(separator)
        
        # Details container with larger font
        details_widget = QWidget()
        details_layout = QVBoxLayout(details_widget)
        details_layout.setSpacing(16)
        details_layout.setContentsMargins(12, 12, 12, 12)
        
        # Card ID
        card_layout = QHBoxLayout()
        card_label = QLabel("Card ID:")
        card_label.setFont(QFont("Arial", 12, QFont.Bold))
        card_label.setMinimumWidth(120)
        card_label.setStyleSheet("color: #212121;")
        card_value = QLabel(card_id)
        card_value.setFont(QFont("Courier New", 13, QFont.Bold))
        card_value.setStyleSheet("color: #1565c0; background-color: #e3f2fd; padding: 8px 12px; border-radius: 4px;")
        card_layout.addWidget(card_label)
        card_layout.addWidget(card_value)
        card_layout.addStretch()
        details_layout.addLayout(card_layout)
        
        # License Plate
        plate_layout = QHBoxLayout()
        plate_label = QLabel("License Plate:")
        plate_label.setFont(QFont("Arial", 12, QFont.Bold))
        plate_label.setMinimumWidth(120)
        plate_label.setStyleSheet("color: #212121;")
        plate_value = QLabel(license_plate)
        plate_value.setFont(QFont("Courier New", 14, QFont.Bold))
        plate_value.setStyleSheet("color: #6d28d9; background-color: #f3e8ff; padding: 8px 12px; border-radius: 4px; letter-spacing: 2px;")
        plate_layout.addWidget(plate_label)
        plate_layout.addWidget(plate_value)
        plate_layout.addStretch()
        details_layout.addLayout(plate_layout)
        
        # Parking Slot
        slot_layout = QHBoxLayout()
        slot_label = QLabel("Parking Slot:")
        slot_label.setFont(QFont("Arial", 12, QFont.Bold))
        slot_label.setMinimumWidth(120)
        slot_label.setStyleSheet("color: #212121;")
        slot_value = QLabel(str(slot))
        slot_value.setFont(QFont("Courier New", 13, QFont.Bold))
        slot_value.setStyleSheet("color: #d32f2f; background-color: #ffebee; padding: 8px 12px; border-radius: 4px;")
        slot_layout.addWidget(slot_label)
        slot_layout.addWidget(slot_value)
        slot_layout.addStretch()
        details_layout.addLayout(slot_layout)
        
        # Time
        time_layout = QHBoxLayout()
        time_label = QLabel("Time:")
        time_label.setFont(QFont("Arial", 12, QFont.Bold))
        time_label.setMinimumWidth(120)
        time_label.setStyleSheet("color: #212121;")
        time_value = QLabel(time_str)
        time_value.setFont(QFont("Courier New", 12, QFont.Bold))
        time_value.setStyleSheet("color: #f57f17; background-color: #fff3e0; padding: 8px 12px; border-radius: 4px;")
        time_layout.addWidget(time_label)
        time_layout.addWidget(time_value)
        time_layout.addStretch()
        details_layout.addLayout(time_layout)
        
        details_widget.setStyleSheet("background-color: white; border-radius: 8px; border: 1px solid #e0e0e0;")
        layout.addWidget(details_widget, 1)
        
        # Question
        question = QLabel("Do you want to proceed with CHECK IN?")
        question.setFont(QFont("Arial", 12, QFont.Bold))
        question.setStyleSheet("color: #1565c0;")
        question.setAlignment(Qt.AlignCenter)
        layout.addWidget(question)
        
        # Buttons
        button_layout = QHBoxLayout()
        button_layout.setSpacing(16)
        
        cancel_btn = QPushButton("✕ CANCEL")
        cancel_btn.setFont(QFont("Arial", 11, QFont.Bold))
        cancel_btn.setMinimumHeight(44)
        cancel_btn.setMinimumWidth(140)
        cancel_btn.setStyleSheet(
            "QPushButton { background-color: #757575; color: white; border-radius: 6px; border: none; }"
            "QPushButton:hover { background-color: #616161; }"
        )
        cancel_btn.clicked.connect(self.reject)
        button_layout.addWidget(cancel_btn)
        
        button_layout.addStretch()
        
        ok_btn = QPushButton("✓ CONFIRM CHECK IN")
        ok_btn.setFont(QFont("Arial", 11, QFont.Bold))
        ok_btn.setMinimumHeight(44)
        ok_btn.setMinimumWidth(200)
        ok_btn.setStyleSheet(
            "QPushButton { background-color: #2e7d32; color: white; border-radius: 6px; border: none; }"
            "QPushButton:hover { background-color: #1b5e20; }"
        )
        ok_btn.clicked.connect(self.accept)
        button_layout.addWidget(ok_btn)
        
        layout.addLayout(button_layout)
        
        self.setLayout(layout)


class CheckOutConfirmationDialog(QDialog):
    """Large, professional check-out confirmation dialog with better UI"""
    
    def __init__(self, parent, license_plate, slot, duration, fee, time_str):
        super().__init__(parent)
        self.setWindowTitle("🚪 CHECK OUT CONFIRMATION")
        self.setGeometry(200, 200, 600, 480)
        self.setStyleSheet("background-color: #f5f5f5;")
        self.user_confirmed = False
        
        self.init_ui(license_plate, slot, duration, fee, time_str)
    
    def init_ui(self, license_plate, slot, duration, fee, time_str):
        """Initialize confirmation dialog UI with large fonts"""
        layout = QVBoxLayout()
        layout.setSpacing(20)
        layout.setContentsMargins(24, 24, 24, 24)
        
        # Title with icon
        title_layout = QHBoxLayout()
        title = QLabel("✓ READY FOR CHECK OUT")
        title.setFont(QFont("Arial", 16, QFont.Bold))
        title.setStyleSheet("color: #c62828;")
        title_layout.addWidget(title)
        title_layout.addStretch()
        layout.addLayout(title_layout)
        
        # Separator line
        separator = QLabel()
        separator.setStyleSheet("background-color: #ef5350; height: 2px;")
        separator.setMinimumHeight(2)
        layout.addWidget(separator)
        
        # Details container with larger font
        details_widget = QWidget()
        details_layout = QVBoxLayout(details_widget)
        details_layout.setSpacing(16)
        details_layout.setContentsMargins(12, 12, 12, 12)
        
        # License Plate
        plate_layout = QHBoxLayout()
        plate_label = QLabel("License Plate:")
        plate_label.setFont(QFont("Arial", 12, QFont.Bold))
        plate_label.setMinimumWidth(140)
        plate_label.setStyleSheet("color: #212121;")
        plate_value = QLabel(license_plate)
        plate_value.setFont(QFont("Courier New", 14, QFont.Bold))
        plate_value.setStyleSheet("color: #6d28d9; background-color: #f3e8ff; padding: 8px 12px; border-radius: 4px; letter-spacing: 2px;")
        plate_layout.addWidget(plate_label)
        plate_layout.addWidget(plate_value)
        plate_layout.addStretch()
        details_layout.addLayout(plate_layout)
        
        # Parking Slot
        slot_layout = QHBoxLayout()
        slot_label = QLabel("Parking Slot:")
        slot_label.setFont(QFont("Arial", 12, QFont.Bold))
        slot_label.setMinimumWidth(140)
        slot_label.setStyleSheet("color: #212121;")
        slot_value = QLabel(str(slot))
        slot_value.setFont(QFont("Courier New", 13, QFont.Bold))
        slot_value.setStyleSheet("color: #d32f2f; background-color: #ffebee; padding: 8px 12px; border-radius: 4px;")
        slot_layout.addWidget(slot_label)
        slot_layout.addWidget(slot_value)
        slot_layout.addStretch()
        details_layout.addLayout(slot_layout)
        
        # Duration
        duration_layout = QHBoxLayout()
        duration_label = QLabel("Duration:")
        duration_label.setFont(QFont("Arial", 12, QFont.Bold))
        duration_label.setMinimumWidth(140)
        duration_label.setStyleSheet("color: #212121;")
        duration_value = QLabel(f"{duration:.2f} hours")
        duration_value.setFont(QFont("Courier New", 12, QFont.Bold))
        duration_value.setStyleSheet("color: #f57f17; background-color: #fff3e0; padding: 8px 12px; border-radius: 4px;")
        duration_layout.addWidget(duration_label)
        duration_layout.addWidget(duration_value)
        duration_layout.addStretch()
        details_layout.addLayout(duration_layout)
        
        # Fee (highlighted in green)
        fee_layout = QHBoxLayout()
        fee_label = QLabel("Parking Fee:")
        fee_label.setFont(QFont("Arial", 12, QFont.Bold))
        fee_label.setMinimumWidth(140)
        fee_label.setStyleSheet("color: #212121;")
        fee_value = QLabel(f"{fee:,} VND")
        fee_value.setFont(QFont("Courier New", 13, QFont.Bold))
        fee_value.setStyleSheet("color: #2e7d32; background-color: #e8f5e9; padding: 8px 12px; border-radius: 4px;")
        fee_layout.addWidget(fee_label)
        fee_layout.addWidget(fee_value)
        fee_layout.addStretch()
        details_layout.addLayout(fee_layout)
        
        # Time
        time_layout = QHBoxLayout()
        time_label = QLabel("Time:")
        time_label.setFont(QFont("Arial", 12, QFont.Bold))
        time_label.setMinimumWidth(140)
        time_label.setStyleSheet("color: #212121;")
        time_value = QLabel(time_str)
        time_value.setFont(QFont("Courier New", 12, QFont.Bold))
        time_value.setStyleSheet("color: #1565c0; background-color: #e3f2fd; padding: 8px 12px; border-radius: 4px;")
        time_layout.addWidget(time_label)
        time_layout.addWidget(time_value)
        time_layout.addStretch()
        details_layout.addLayout(time_layout)
        
        details_widget.setStyleSheet("background-color: white; border-radius: 8px; border: 1px solid #e0e0e0;")
        layout.addWidget(details_widget, 1)
        
        # Question
        question = QLabel("Do you want to proceed with CHECK OUT?")
        question.setFont(QFont("Arial", 12, QFont.Bold))
        question.setStyleSheet("color: #c62828;")
        question.setAlignment(Qt.AlignCenter)
        layout.addWidget(question)
        
        # Buttons
        button_layout = QHBoxLayout()
        button_layout.setSpacing(16)
        
        cancel_btn = QPushButton("✕ CANCEL")
        cancel_btn.setFont(QFont("Arial", 11, QFont.Bold))
        cancel_btn.setMinimumHeight(44)
        cancel_btn.setMinimumWidth(140)
        cancel_btn.setStyleSheet(
            "QPushButton { background-color: #757575; color: white; border-radius: 6px; border: none; }"
            "QPushButton:hover { background-color: #616161; }"
        )
        cancel_btn.clicked.connect(self.reject)
        button_layout.addWidget(cancel_btn)
        
        button_layout.addStretch()
        
        ok_btn = QPushButton("✓ CONFIRM CHECK OUT")
        ok_btn.setFont(QFont("Arial", 11, QFont.Bold))
        ok_btn.setMinimumHeight(44)
        ok_btn.setMinimumWidth(200)
        ok_btn.setStyleSheet(
            "QPushButton { background-color: #c62828; color: white; border-radius: 6px; border: none; }"
            "QPushButton:hover { background-color: #ad1457; }"
        )
        ok_btn.clicked.connect(self.accept)
        button_layout.addWidget(ok_btn)
        
        layout.addLayout(button_layout)
        
        self.setLayout(layout)


class ParkingManagementUI(QMainWindow):
    """Main window for Parking Management System with YOLO detection and serial port support"""
    
    def __init__(self):
        super().__init__()
        self.setWindowTitle("🅿️ Parking Management System")
        self.setGeometry(50, 50, 1920, 1080)
        
        # Color scheme for user-friendly parking management system
        self.PRIMARY_COLOR = "#1E88E5"      # Modern friendly blue
        self.PRIMARY_LIGHT = "#42A5F5"      # Lighter variant for hover
        self.PRIMARY_DARK = "#1565C0"       # Darker variant for active
        self.SECONDARY_COLOR = "#26A69A"    # Teal for capture actions
        self.SUCCESS_COLOR = "#43A047"      # Friendly green for positive
        self.DANGER_COLOR = "#E53935"       # Soft red for negative
        self.NEUTRAL_BG = "#ECEFF1"         # Light cool background
        
        # Data storage
        self.parking_records = []
        self.models_path = Path(__file__).parent / "Automatic-License-Plate-Recognition-using-YOLOv8"
        self.csv_file = Path(__file__).parent / "parking_records.csv"
        self.current_frame = None
        self.captured_license_plate = None  # Store extracted license plate from OCR
        
        # Temporary variables for serial card processing
        self.pending_card_data = None  # Store card data while waiting for OCR
        self.is_processing_card = False  # Flag to prevent concurrent card processing
        
        # UI placeholder references (will be initialized in init_ui)
        self.status_label = None
        self.fee_label = None
        
        # Initialize CSV file
        self._init_csv()
        
        # Camera Signals
        self.signals = CameraSignals()
        self.signals.frame_ready.connect(self.on_frame_received)
        self.signals.error_signal.connect(self.on_camera_error)
        
        # Serial Signals
        self.serial_signals = SerialSignals()
        self.serial_signals.data_received.connect(self.on_card_scanned)
        self.serial_signals.connection_status.connect(self.on_serial_status_changed)
        self.serial_signals.error_signal.connect(self.on_serial_error)
        
        # Serial worker thread
        self.serial_thread = None
        self.serial_connected = False
        
        # Build user interface
        self.init_ui()
        
        # Start camera worker thread
        self.camera_thread = CameraWorker(self.signals, self.models_path)
        self.camera_thread.start()
        
        # Startup timer for initialization message
        self.start_timer = QTimer()
        self.start_timer.setSingleShot(True)
        self.start_timer.timeout.connect(lambda: print("✓ System ready"))
    
    def _init_csv(self):
        """Initialize CSV file with headers if it doesn't exist"""
        if not self.csv_file.exists():
            with open(self.csv_file, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(['Card ID', 'License Plate', 'Time In', 'Time Out', 'Slot Number', 'Status', 'Fee (VND)'])
    
    def init_ui(self):
        """Initialize user interface"""
        central = QWidget()
        central.setStyleSheet("background-color: #f5f6fa;")
        self.setCentralWidget(central)
        
        # ===== MAIN LAYOUT =====
        main_layout = QHBoxLayout(central)
        main_layout.setSpacing(12)
        main_layout.setContentsMargins(12, 12, 12, 12)
        
        # ====================== LEFT PANEL (Smaller, Top-Left Video + Bottom Snapshot) ======================
        left_layout = QVBoxLayout()
        left_layout.setSpacing(8)
        left_layout.setContentsMargins(0, 0, 0, 0)
        left_panel = QWidget()
        left_panel.setLayout(left_layout)
        left_panel.setMaximumWidth(620)
        left_panel.setStyleSheet("background-color: white; border-radius: 10px; border: 2px solid #e0e0e0;")
        
        # ---- VIDEO FEED SECTION (Top Left) ----
        video_container = QWidget()
        video_layout = QVBoxLayout(video_container)
        video_layout.setSpacing(6)
        video_layout.setContentsMargins(12, 12, 12, 12)
        video_container.setStyleSheet("background-color: white;")
        
        # Video header with icon
        video_header = QLabel("CAMERA")
        video_header.setFont(QFont("Arial", 10, QFont.Bold))
        video_header.setAlignment(Qt.AlignCenter)
        video_header.setStyleSheet(
            "background-color: #1e3a8a; color: white; padding: 8px; border-radius: 5px; "
            "font-weight: bold; letter-spacing: 1px;"
        )
        video_layout.addWidget(video_header)
        
        # Camera label (Compact size - 400x280)
        self.camera_label = QLabel()
        self.camera_label.setAlignment(Qt.AlignCenter)
        self.camera_label.setMinimumSize(480, 320)
        self.camera_label.setMaximumSize(580, 380)
        self.camera_label.setStyleSheet(
            "border: 3px solid #1e3a8a; background-color: #000000; border-radius: 8px; padding: 0px;"
        )
        video_layout.addWidget(self.camera_label, alignment=Qt.AlignCenter)
        video_layout.addStretch()
        
        left_layout.addWidget(video_container)
        
        # ---- SNAPSHOT SECTION (Bottom) ----
        snapshot_container = QWidget()
        snapshot_layout = QVBoxLayout(snapshot_container)
        snapshot_layout.setSpacing(6)
        snapshot_layout.setContentsMargins(12, 12, 12, 12)
        snapshot_container.setStyleSheet("background-color: white;")
        
        # Snapshot header
        snapshot_header = QLabel("LICENSE PLATE IMAGE")
        snapshot_header.setFont(QFont("Arial", 10, QFont.Bold))
        snapshot_header.setAlignment(Qt.AlignCenter)
        snapshot_header.setStyleSheet(
            "background-color: #059669; color: white; padding: 8px; border-radius: 5px; "
            "font-weight: bold; letter-spacing: 1px;"
        )
        snapshot_layout.addWidget(snapshot_header)
        
        # Snapshot display label
        self.snapshot_label = QLabel()
        self.snapshot_label.setAlignment(Qt.AlignCenter)
        self.snapshot_label.setMinimumSize(480, 120)
        self.snapshot_label.setMaximumSize(580, 160)
        self.snapshot_label.setStyleSheet(
            "border: 2px dashed #059669; background-color: #ecfdf5; border-radius: 6px; "
            "color: #047857; font-weight: bold;"
        )
        self.snapshot_label.setText("No snapshot captured yet\nClick 'CAPTURE LICENSE PLATE' to start")
        snapshot_layout.addWidget(self.snapshot_label, alignment=Qt.AlignCenter)
        
        # Extracted license plate label (Large, Bold)
        plate_info_header = QLabel("EXTRACTED TEXT")
        plate_info_header.setFont(QFont("Arial", 9, QFont.Bold))
        plate_info_header.setAlignment(Qt.AlignCenter)
        plate_info_header.setStyleSheet(
            "background-color: #7c3aed; color: white; padding: 6px; border-radius: 4px; margin-top: 8px;"
        )
        snapshot_layout.addWidget(plate_info_header)
        
        self.extracted_plate_label = QLabel("Awaiting capture...")
        self.extracted_plate_label.setAlignment(Qt.AlignCenter)
        self.extracted_plate_label.setFont(QFont("Courier New", 16, QFont.Bold))
        self.extracted_plate_label.setStyleSheet(
            "background-color: #f3e8ff; color: #6d28d9; padding: 10px; "
            "border-radius: 6px; border: 2px solid #7c3aed; letter-spacing: 2px;"
        )
        snapshot_layout.addWidget(self.extracted_plate_label)
        
        left_layout.addWidget(snapshot_container)
        
        # ---- PARKING SLOTS SECTION (Bottom) ----
        slots_container = QWidget()
        slots_layout = QVBoxLayout(slots_container)
        slots_layout.setSpacing(6)
        slots_layout.setContentsMargins(12, 12, 12, 12)
        slots_container.setStyleSheet("background-color: white;")
        
        # Slots header
        slots_header = QLabel("PARKING PALLET STATUS")
        slots_header.setFont(QFont("Arial", 10, QFont.Bold))
        slots_header.setAlignment(Qt.AlignCenter)
        slots_header.setStyleSheet(
            "background-color: #1f2937; color: white; padding: 8px; border-radius: 5px; "
            "font-weight: bold; letter-spacing: 1px;"
        )
        slots_layout.addWidget(slots_header)
        
        # Grid of parking slots (2 rows x 5 columns = 10 slots)
        slots_grid = QWidget()
        grid_layout = QVBoxLayout(slots_grid)
        grid_layout.setSpacing(8)
        grid_layout.setContentsMargins(8, 8, 8, 8)
        
        # Initialize slot buttons/labels (10 slots)
        self.slot_buttons = []
        
        for row in range(2):
            row_widget = QWidget()
            row_layout = QHBoxLayout(row_widget)
            row_layout.setSpacing(8)
            row_layout.setContentsMargins(0, 0, 0, 0)
            
            for col in range(5):
                slot_num = row * 5 + col + 1  # Slot 1-10
                
                slot_btn = QPushButton(f"PALLET {slot_num}")
                slot_btn.setMinimumHeight(50)
                slot_btn.setMinimumWidth(80)  # Ensure equal width
                slot_btn.setMaximumWidth(120)  # Cap maximum width for consistency
                slot_btn.setFont(QFont("Arial", 9, QFont.Bold))
                slot_btn.setStyleSheet(
                    "QPushButton {"
                    "  background-color: #10b981;"
                    "  color: white;"
                    "  font-weight: bold;"
                    "  border-radius: 6px;"
                    "  border: 2px solid #059669;"
                    "  padding: 6px;"
                    "}"
                    "QPushButton:hover {"
                    "  background-color: #059669;"
                    "}"
                )
                slot_btn.setEnabled(False)  # Disable clicking
                
                self.slot_buttons.append({
                    'button': slot_btn,
                    'slot_num': slot_num,
                    'is_occupied': False
                })
                
                row_layout.addWidget(slot_btn, alignment=Qt.AlignCenter)
            
            grid_layout.addWidget(row_widget)
        
        # Add HOME button below parking slots
        home_btn_layout = QHBoxLayout()
        home_btn_layout.addStretch()
        
        self.home_btn = QPushButton("🏠 HOME")
        self.home_btn.setMinimumHeight(50)
        self.home_btn.setMinimumWidth(120)
        self.home_btn.setMaximumWidth(160)
        self.home_btn.setFont(QFont("Arial", 11, QFont.Bold))
        self.home_btn.setStyleSheet(
            "QPushButton {"
            "  background-color: #0284c7;"
            "  color: white;"
            "  font-weight: bold;"
            "  border-radius: 6px;"
            "  border: 2px solid #0369a1;"
            "  padding: 8px;"
            "  letter-spacing: 0.5px;"
            "}"
            "QPushButton:hover {"
            "  background-color: #0369a1;"
            "}"
            "QPushButton:pressed {"
            "  background-color: #075985;"
            "}"
        )
        self.home_btn.clicked.connect(self.on_home_button_clicked)
        
        home_btn_layout.addWidget(self.home_btn, alignment=Qt.AlignCenter)
        home_btn_layout.addStretch()
        
        grid_layout.addLayout(home_btn_layout)
        
        slots_grid.setStyleSheet("background-color: white;")
        slots_layout.addWidget(slots_grid)
        
        left_layout.addWidget(slots_container)
        left_layout.addStretch()
        
        # ====================== RIGHT PANEL (Professional Controls) ======================
        right_layout = QVBoxLayout()
        right_layout.setSpacing(10)
        right_layout.setContentsMargins(0, 0, 0, 0)
        right_panel = QWidget()
        right_panel.setLayout(right_layout)
        right_panel.setStyleSheet("background-color: white; border-radius: 10px; border: 2px solid #e0e0e0;")
        
        # Scroll area for better organization
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setStyleSheet("QScrollArea { border: none; background-color: white; }")
        
        scroll_widget = QWidget()
        scroll_layout = QVBoxLayout(scroll_widget)
        scroll_layout.setSpacing(10)
        scroll_layout.setContentsMargins(12, 12, 12, 12)
        
        # ============ SECTION 0: SERIAL PORT CONFIGURATION (Material Design) ============
        # Settings section header (Title)
        # settings_header = QLabel("SETTINGS & CONNECTION")
        # settings_header.setFont(QFont("Arial", 11, QFont.Bold))
        # settings_header.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        # settings_header.setStyleSheet(
        #     f"background-color: {self.PRIMARY_DARK}; color: white; padding: 10px 12px; border-radius: 4px; "
        #     "font-weight: bold; letter-spacing: 0.5px;"
        # )
        # scroll_layout.addWidget(settings_header)
        
        # Serial port settings container
        serial_section = QWidget()
        serial_layout = QVBoxLayout(serial_section)
        serial_layout.setSpacing(12)
        serial_layout.setContentsMargins(12, 12, 12, 12)
        
        # Connection status indicator and CONNECT button
        status_layout = QHBoxLayout()
        status_layout.setSpacing(12)
        
        self.serial_status_label = QLabel("🔴 Disconnected")
        self.serial_status_label.setFont(QFont("Arial", 10, QFont.Bold))
        self.serial_status_label.setStyleSheet("color: #d32f2f;")
        status_layout.addWidget(self.serial_status_label)
        status_layout.addStretch()
        
        self.serial_connect_btn = self._create_button(
            "⚙️ SETTING",
            self.PRIMARY_COLOR,
            self.PRIMARY_LIGHT,
            36,
            self.open_settings_dialog
        )
        self.serial_connect_btn.setMaximumWidth(130)
        status_layout.addWidget(self.serial_connect_btn)
        
        serial_layout.addLayout(status_layout)
        
        serial_section.setStyleSheet(
            "background-color: #f0f4f8; border-radius: 8px; border: 1px solid #e5e7eb;"
        )
        scroll_layout.addWidget(serial_section)
        
        # ============ SECTION 1: CHECK IN (Vehicle Entry) ============
        checkin_header = QLabel("CHECK IN")
        checkin_header.setFont(QFont("Arial", 11, QFont.Bold))
        checkin_header.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        checkin_header.setStyleSheet(
            f"background-color: {self.PRIMARY_DARK}; color: white; padding: 10px 12px; border-radius: 4px; "
            "font-weight: bold; letter-spacing: 0.5px;"
        )
        scroll_layout.addWidget(checkin_header)
        
        # CHECK IN container
        checkin_section = QWidget()
        checkin_layout = QVBoxLayout(checkin_section)
        checkin_layout.setSpacing(8)
        checkin_layout.setContentsMargins(12, 12, 12, 12)
        
        # Row 1: Card ID input + Capture button
        row1_layout = QHBoxLayout()
        row1_layout.setSpacing(8)
        
        card_label = QLabel("Card ID")
        card_label.setFont(QFont("Arial", 10, QFont.Bold))
        card_label.setMinimumWidth(80)
        card_label.setStyleSheet("color: #212121;")
        row1_layout.addWidget(card_label)
        
        self.card_id_input = QLineEdit()
        self.card_id_input.setPlaceholderText("Enter card ID...")
        self.card_id_input.setMinimumHeight(36)
        self.card_id_input.setStyleSheet(
            "border: 1px solid #bdbdbd; border-radius: 4px; padding: 8px; background-color: #ffffff;"
            "QLineEdit:focus { border: 2px solid #2196F3; padding: 7px; }"
        )
        row1_layout.addWidget(self.card_id_input)
        
        self.capture_btn = self._create_button(
            "📷 CAPTURE",
            self.SECONDARY_COLOR,
            "#1b7470",
            36,
            self.capture_plate
        )
        row1_layout.addWidget(self.capture_btn)
        
        checkin_layout.addLayout(row1_layout)
        
        # Row 2: Slot Number input + Check-in button
        row2_layout = QHBoxLayout()
        row2_layout.setSpacing(8)
        
        slot_label = QLabel("Slot")
        slot_label.setFont(QFont("Arial", 10, QFont.Bold))
        slot_label.setMinimumWidth(80)
        slot_label.setStyleSheet("color: #212121;")
        row2_layout.addWidget(slot_label)
        
        self.slot_input = QLineEdit()
        self.slot_input.setPlaceholderText("Enter slot number...")
        self.slot_input.setMinimumHeight(36)
        self.slot_input.setStyleSheet(
            "border: 1px solid #bdbdbd; border-radius: 4px; padding: 8px; background-color: #ffffff;"
            "QLineEdit:focus { border: 2px solid #2196F3; padding: 7px; }"
        )
        row2_layout.addWidget(self.slot_input)
        
        self.checkin_btn = self._create_button(
            "✅ CHECK IN",
            self.SUCCESS_COLOR,
            "#2e7d32",
            36,
            self.check_in
        )
        row2_layout.addWidget(self.checkin_btn)
        
        checkin_layout.addLayout(row2_layout)
        
        # Row 3: Time In Display
        time_in_layout = QHBoxLayout()
        time_in_layout.setSpacing(8)
        
        time_in_label_text = QLabel("Time In")
        time_in_label_text.setFont(QFont("Arial", 10, QFont.Bold))
        time_in_label_text.setMinimumWidth(80)
        time_in_label_text.setStyleSheet("color: #212121;")
        time_in_layout.addWidget(time_in_label_text)
        
        self.time_in_display = QLabel("--:--:--")
        self.time_in_display.setFont(QFont("Courier New", 11, QFont.Bold))
        self.time_in_display.setMinimumHeight(36)
        self.time_in_display.setStyleSheet(
            "background-color: #e3f2fd; color: #1565c0; padding: 8px; border-radius: 4px; "
            "border: 1px solid #90caf9;"
        )
        time_in_layout.addWidget(self.time_in_display)
        time_in_layout.addStretch()
        
        checkin_layout.addLayout(time_in_layout)
        
        checkin_section.setStyleSheet(
            "background-color: #f0f4f8; border-radius: 8px; border: 1px solid #e5e7eb;"
        )
        scroll_layout.addWidget(checkin_section)
        
        # ============ SECTION 2: CHECK OUT (Vehicle Exit) ============
        checkout_header = QLabel("CHECK OUT")
        checkout_header.setFont(QFont("Arial", 11, QFont.Bold))
        checkout_header.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        checkout_header.setStyleSheet(
            f"background-color: {self.PRIMARY_DARK}; color: white; padding: 10px 12px; border-radius: 4px; "
            "font-weight: bold; letter-spacing: 0.5px;"
        )
        scroll_layout.addWidget(checkout_header)
        
        # CHECK OUT container
        checkout_section = QWidget()
        checkout_layout = QVBoxLayout(checkout_section)
        checkout_layout.setSpacing(8)
        checkout_layout.setContentsMargins(12, 12, 12, 12)
        
        # License Plate input + Check-out button
        plate_layout = QHBoxLayout()
        plate_layout.setSpacing(8)
        
        plate_label = QLabel("Plate")
        plate_label.setFont(QFont("Arial", 10, QFont.Bold))
        plate_label.setMinimumWidth(80)
        plate_label.setStyleSheet("color: #212121;")
        plate_layout.addWidget(plate_label)
        
        self.checkout_plate_input = QLineEdit()
        self.checkout_plate_input.setPlaceholderText("Enter license plate...")
        self.checkout_plate_input.setMinimumHeight(36)
        self.checkout_plate_input.setStyleSheet(
            "border: 1px solid #bdbdbd; border-radius: 4px; padding: 8px; background-color: #ffffff;"
            "QLineEdit:focus { border: 2px solid #2196F3; padding: 7px; }"
        )
        plate_layout.addWidget(self.checkout_plate_input)
        
        self.checkout_btn = self._create_button(
            "🏁 CHECK OUT",
            self.DANGER_COLOR,
            "#c62828",
            36,
            self.check_out
        )
        plate_layout.addWidget(self.checkout_btn)
        
        checkout_layout.addLayout(plate_layout)
        
        # Time Out Display
        time_out_layout = QHBoxLayout()
        time_out_layout.setSpacing(8)
        
        time_out_label_text = QLabel("Time Out")
        time_out_label_text.setFont(QFont("Arial", 10, QFont.Bold))
        time_out_label_text.setMinimumWidth(80)
        time_out_label_text.setStyleSheet("color: #212121;")
        time_out_layout.addWidget(time_out_label_text)
        
        self.time_out_display = QLabel("--:--:--")
        self.time_out_display.setFont(QFont("Courier New", 11, QFont.Bold))
        self.time_out_display.setMinimumHeight(36)
        self.time_out_display.setStyleSheet(
            "background-color: #ffe0b2; color: #e65100; padding: 8px; border-radius: 4px; "
            "border: 1px solid #ffb74d;"
        )
        time_out_layout.addWidget(self.time_out_display)
        time_out_layout.addStretch()
        
        checkout_layout.addLayout(time_out_layout)
        
        checkout_section.setStyleSheet(
            "background-color: #f0f4f8; border-radius: 8px; border: 1px solid #e5e7eb;"
        )
        scroll_layout.addWidget(checkout_section)
        
        # ============ SECTION 3: PARKING RECORDS TABLE ============
        table_header = QLabel("PARKING RECORDS & HISTORY")
        table_header.setFont(QFont("Arial", 11, QFont.Bold))
        table_header.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        table_header.setStyleSheet(
            f"background-color: {self.PRIMARY_DARK}; color: white; padding: 10px 12px; border-radius: 4px; "
            "font-weight: bold; letter-spacing: 0.5px;"
        )
        scroll_layout.addWidget(table_header)
        
        # Table with Material Design styling
        self.table = QTableWidget()
        self.table.setColumnCount(7)
        self.table.setHorizontalHeaderLabels([
            'Card ID', 'License Plate', 'Time In', 'Time Out', 'Slot', 'Status', 'Fee (VND)'
        ])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.setRowCount(0)
        self.table.setMinimumHeight(200)
        self.table.setAlternatingRowColors(True)
        self.table.setStyleSheet(f"""
            QTableWidget {{
                background-color: white;
                alternate-background-color: #f5f5f5;
                border-radius: 4px;
                border: 1px solid #e0e0e0;
            }}
            QTableWidget::item {{
                padding: 8px;
            }}
            QHeaderView::section {{
                background-color: {self.PRIMARY_COLOR};
                color: white;
                padding: 10px;
                border: none;
                font-weight: bold;
                border-radius: 0px;
            }}
        """)
        scroll_layout.addWidget(self.table, 1)
        
        # Download Excel button - centered
        download_layout = QHBoxLayout()
        download_layout.addStretch()  # Add space on left
        
        self.download_btn = self._create_button(
            "📥 REPORT",
            self.SUCCESS_COLOR,
            "#38764d",
            44,
            self.export_to_excel
        )
        download_layout.addWidget(self.download_btn)
        download_layout.addStretch()  # Add space on right
        
        scroll_layout.addLayout(download_layout)
        
        scroll_layout.addStretch()
        
        # Add scroll widget to scroll area
        scroll_area.setWidget(scroll_widget)
        right_layout.addWidget(scroll_area)
        
        # Add panels to main layout
        main_layout.addWidget(left_panel, 0)
        main_layout.addWidget(right_panel, 1)
    
    def _create_button(self, text, bg_color, hover_color, height, callback):
        """Create a professional button with consistent styling and fixed width"""
        btn = QPushButton(text)
        btn.setMinimumHeight(height)
        
        # Special handling for Download button
        if "DOWNLOAD" in text:
            btn.setMinimumWidth(180)
            btn.setMaximumWidth(220)
        else:
            btn.setMinimumWidth(125)
            btn.setMaximumWidth(145)
        
        btn.setFont(QFont("Arial", 9, QFont.Bold))
        btn.setStyleSheet(
            f"QPushButton {{"
            f"  background-color: {bg_color};"
            f"  color: white;"
            f"  font-weight: bold;"
            f"  border-radius: 6px;"
            f"  border: none;"
            f"  padding: 8px;"
            f"  letter-spacing: 0.5px;"
            f"}}"
            f"QPushButton:hover {{"
            f"  background-color: {hover_color};"
            f"}}"
            f"QPushButton:pressed {{"
            f"  background-color: {hover_color};"
            f"  padding: 10px 8px 6px 8px;"
            f"}}"
        )
        btn.clicked.connect(callback)
        return btn
    
    def _update_status(self, message):
        """Safely update status label if it exists"""
        try:
            if self.status_label:
                self.status_label.setText(message)
        except Exception as e:
            print(f"⚠️  Error updating status: {e}")
    
    def _refresh_ports(self):
        """Auto-detect and populate available serial ports"""
        self.port_combo.clear()
        ports = get_available_ports()
        
        if ports:
            for port_name, port_desc in ports:
                self.port_combo.addItem(f"{port_name} - {port_desc}", port_name)
        else:
            self.port_combo.addItem("No ports found", None)
    
    def toggle_serial_connection(self):
        """Toggle serial port connection - Connect or disconnect from Arduino/ESP32"""
        if self.serial_connected:
            # Disconnect from serial port
            if self.serial_thread:
                self.serial_thread.stop()
                self.serial_thread.join(timeout=2)
                self.serial_thread = None
            self.serial_connected = False
            self.on_serial_status_changed(False, "Disconnected")
        else:
            # Connect to selected serial port
            port_name = self.port_combo.currentData()
            baud_rate = int(self.baud_combo.currentText())
            
            if not port_name:
                QMessageBox.warning(self, "Error", "No serial port selected")
                return
            
            # Create and start serial worker thread
            self.serial_thread = SerialWorker(self.serial_signals)
            self.serial_thread.set_connection(port_name, baud_rate)
            self.serial_thread.start()
            self.serial_connected = True
    
    def open_settings_dialog(self):
        """Open serial port settings dialog (Material Design)"""
        try:
            dialog = SettingsDialog(self)
            if dialog.exec_() == QDialog.Accepted:
                settings = dialog.get_settings()
                if settings['port']:
                    try:
                        # Disconnect if connected
                        if self.serial_connected and self.serial_thread:
                            try:
                                self.serial_thread.stop()
                                self.serial_thread.join(timeout=2)
                                self.serial_thread = None
                                self.serial_connected = False
                            except Exception as e:
                                print(f"⚠️  Error disconnecting previous connection: {e}")
                                self.serial_thread = None
                                self.serial_connected = False
                        
                        # Connect with new settings
                        self.serial_thread = SerialWorker(self.serial_signals)
                        self.serial_thread.set_connection(settings['port'], settings['baud'])
                        self.serial_thread.start()
                        self.serial_connected = True
                        print(f"✓ Connected to {settings['port']} @ {settings['baud']} baud")
                    except Exception as e:
                        print(f"❌ Connection error: {e}")
                        self.serial_connected = False
                        self.serial_thread = None
                        QMessageBox.warning(
                            self, 
                            "Connection Failed", 
                            f"Failed to connect to serial port:\n\n{str(e)}"
                        )
        except Exception as e:
            print(f"❌ Settings dialog error: {e}")
            QMessageBox.critical(
                self, 
                "Error", 
                f"An error occurred in settings dialog:\n\n{str(e)}"
            )
    
    def on_serial_status_changed(self, connected, message):
        """Update serial connection status indicator in UI"""
        self.serial_connected = connected
        
        if connected:
            # Show connected status (green indicator)
            self.serial_status_label.setText("🟢 Connected")
            self.serial_status_label.setStyleSheet("color: #388E3C; font-weight: bold; font-size: 11pt;")
            print(f"✓ {message}")
        else:
            # Show disconnected status (red indicator)
            self.serial_status_label.setText("🔴 Disconnected")
            self.serial_status_label.setStyleSheet("color: #D32F2F; font-weight: bold; font-size: 11pt;")
            if message != "Disconnected":
                print(f"✗ {message}")
    
    def on_serial_error(self, error_msg):
        """Handle and display serial port errors safely"""
        try:
            print(f"❌ Serial Error: {error_msg}")
            # Safely update status
            self._update_status(f"Status: Serial Error - {error_msg}")
        except Exception as e:
            print(f"⚠️  Error handling serial error: {e}")
    
    def on_card_scanned(self, card_data):
        """Handle card scan event from Arduino/ESP32 RFID reader
        
        Expected JSON format:
        {"event":"CARD_SCAN","uid":"A1B2C3D4","gate":"ENTRY","slot":1}
        
        Process:
        1. Store card data temporarily
        2. Trigger license plate capture
        3. Wait for OCR extraction to complete
        4. Auto check-in/out based on gate field
        """
        try:
            # Prevent concurrent card processing
            if self.is_processing_card:
                print("⚠️  Already processing a card scan, skipping...")
                return
            
            self.is_processing_card = True
            
            # Extract card data from JSON
            card_id = card_data.get('uid', '').strip()
            gate = card_data.get('gate', 'ENTRY').upper()  # ENTRY or EXIT
            slot = card_data.get('slot', '')
            
            # Validate card data
            if not card_id:
                self._update_status("Status: Invalid card ID received")
                self.is_processing_card = False
                return
            
            # Store card data for later use
            self.pending_card_data = {
                'card_id': card_id,
                'gate': gate,
                'slot': str(slot) if slot else ''
            }
            
            # Auto-fill the card ID and slot fields
            self.card_id_input.setText(card_id)
            if self.pending_card_data['slot']:
                self.slot_input.setText(self.pending_card_data['slot'])
            
            # Update status to show card received
            self._update_status(f"Status: Card scanned ({card_id}) - {gate} | Capturing license plate...")
            print(f"\n📱 Card event received: ID={card_id}, Gate={gate}, Slot={slot}")
            
            # Trigger license plate capture after small delay
            QTimer.singleShot(200, self._process_card_capture)
        
        except Exception as e:
            self._update_status(f"Status: Error processing card data - {str(e)}")
            print(f"❌ Card processing error: {e}")
            self.is_processing_card = False
    
    def _process_card_capture(self):
        """Capture license plate for the pending card scan"""
        if not self.pending_card_data or self.current_frame is None:
            self._update_status("Status: No frame available for capture")
            self.is_processing_card = False
            return
        
        try:
            # Check if YOLO model is loaded
            if not self.camera_thread.license_plate_detector:
                self._update_status("Status: YOLO model not ready")
                self.is_processing_card = False
                return
            
            # Run YOLO detection on current frame
            detections = self.camera_thread.license_plate_detector(self.current_frame)[0]
            
            if len(detections.boxes) == 0:
                self._update_status(f"Status: No license plate detected for {self.pending_card_data['card_id']}")
                print(f"⚠️  No plate detected for card {self.pending_card_data['card_id']}")
                self.is_processing_card = False
                return
            
            # Extract first detection bounding box
            detection = detections.boxes.data.tolist()[0]
            x1, y1, x2, y2, score, class_id = detection
            
            # Crop license plate region
            crop = self.current_frame[int(y1):int(y2), int(x1):int(x2), :]
            
            # Display snapshot (cropped license plate)
            rgb_crop = cv2.cvtColor(crop, cv2.COLOR_BGR2RGB)
            h, w, ch = rgb_crop.shape
            bytes_per_line = 3 * w
            qt_image = QImage(rgb_crop.data, w, h, bytes_per_line, QImage.Format_RGB888)
            pixmap = QPixmap.fromImage(qt_image)
            scaled_pixmap = pixmap.scaledToWidth(380, Qt.SmoothTransformation)
            self.snapshot_label.setPixmap(scaled_pixmap)
            
            # Save captured images to disk
            try:
                captured_dir = Path(__file__).parent / "captured_images"
                captured_dir.mkdir(exist_ok=True)
                
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")[:-3]
                
                # Save full frame
                full_frame_path = captured_dir / f"full_frame_{timestamp}.jpg"
                cv2.imwrite(str(full_frame_path), self.current_frame)
                
                # Save cropped license plate
                crop_path = captured_dir / f"license_plate_{timestamp}.jpg"
                cv2.imwrite(str(crop_path), crop)
                
                print(f"✓ Captured images saved: {timestamp}")
            except Exception as save_error:
                print(f"⚠️  Could not save images: {save_error}")
            
            # Run OCR to extract license plate text
            self._update_status(f"Status: Extracting text for {self.pending_card_data['card_id']}...")
            plate_text, conf = extract_license_plate(crop)
            
            if not plate_text:
                self._update_status(f"Status: OCR failed for {self.pending_card_data['card_id']} - no text extracted")
                self.extracted_plate_label.setText("OCR Failed")
                print(f"❌ OCR extraction failed for card {self.pending_card_data['card_id']}")
                self.is_processing_card = False
                return
            
            # Successfully extracted license plate
            self.captured_license_plate = plate_text
            self.extracted_plate_label.setText(plate_text)
            self.extracted_plate_label.setStyleSheet(
                "background-color: #f3e8ff; color: #6d28d9; padding: 10px; "
                "border-radius: 6px; border: 2px solid #7c3aed; letter-spacing: 2px; "
                "font-size: 16pt; font-weight: bold;"
            )
            
            print(f"✅ License plate extracted: {plate_text} (Confidence: {conf:.2%})")
            
            # Update Time In display for RFID scan
            current_time = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
            self.time_in_display.setText(current_time)
            print(f"   Time In: {current_time}")
            
            # Now perform auto check-in or check-out based on gate
            gate = self.pending_card_data['gate']
            
            # Show confirmation dialog before proceeding
            if gate == "ENTRY":
                self._update_status(f"Status: Waiting for confirmation...")
                QTimer.singleShot(300, self._show_checkin_confirmation)
            elif gate == "EXIT":
                self._update_status(f"Status: Waiting for confirmation...")
                QTimer.singleShot(300, self._show_checkout_confirmation)
            else:
                self._update_status(f"Status: Unknown gate type: {gate}")
                self.is_processing_card = False
        
        except Exception as e:
            self._update_status(f"Status: Capture error - {str(e)}")
            print(f"❌ Capture error: {e}")
            self.is_processing_card = False
    
    def _show_checkin_confirmation(self):
        """Show confirmation dialog for check-in from RFID card"""
        if not self.pending_card_data or not self.captured_license_plate:
            self.is_processing_card = False
            return
        
        card_id = self.pending_card_data['card_id']
        slot = self.pending_card_data['slot']
        plate = self.captured_license_plate
        time_str = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        
        # Show custom confirmation dialog
        dialog = CheckInConfirmationDialog(self, card_id, plate, slot, time_str)
        
        if dialog.exec_() == QDialog.Accepted:
            # Proceed with check-in
            QTimer.singleShot(100, self._auto_check_in)
        else:
            # User cancelled
            self._update_status("Status: Check-in cancelled by user")
            print(f"⚠️  Check-in cancelled for card {card_id}")
            self.pending_card_data = None
            self.is_processing_card = False
    
    def _show_checkout_confirmation(self):
        """Show confirmation dialog for check-out from RFID card"""
        if not self.pending_card_data or not self.captured_license_plate:
            self.is_processing_card = False
            return
        
        card_id = self.pending_card_data['card_id']
        plate = self.captured_license_plate
        
        # Find the record to show duration
        record = None
        for r in self.parking_records:
            if r['license_plate'] == plate and r['status'] == 'IN':
                record = r
                break
        
        duration = 0
        if record:
            time_in = record['time_in']
            time_out = datetime.now()
            duration = (time_out - time_in).total_seconds() / 3600
        
        # Calculate fee
        if duration <= 2:
            fee = 20000
        else:
            fee = 20000 + int((duration - 2) * 10000)
        
        slot = record['slot'] if record else "N/A"
        time_str = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        
        # Show custom confirmation dialog
        dialog = CheckOutConfirmationDialog(self, plate, slot, duration, fee, time_str)
        
        if dialog.exec_() == QDialog.Accepted:
            # Proceed with check-out
            QTimer.singleShot(100, self._auto_check_out)
        else:
            # User cancelled
            self._update_status("Status: Check-out cancelled by user")
            print(f"⚠️  Check-out cancelled for card {card_id}")
            self.pending_card_data = None
            self.is_processing_card = False
    
    def _auto_check_in(self):
        """Auto check-in after successful license plate extraction"""
        try:
            if not self.pending_card_data:
                self.is_processing_card = False
                return
            
            card_id = self.pending_card_data['card_id']
            slot = self.pending_card_data['slot']
            
            if not card_id or not slot:
                self._update_status("Status: Missing card ID or slot for check-in")
                self.is_processing_card = False
                return
            
            # Check if already checked in
            for record in self.parking_records:
                if record['card_id'] == card_id and record['status'] == 'IN':
                    self._update_status(f"Status: {card_id} already checked in at slot {record['slot']}")
                    print(f"⚠️  Card {card_id} already checked in")
                    self.is_processing_card = False
                    return
            
            now = datetime.now()
            
            # Create new parking record
            record = {
                'card_id': card_id,
                'license_plate': self.captured_license_plate,  # Use extracted plate
                'time_in': now,
                'time_out': None,
                'slot': slot,
                'status': 'IN',
                'fee': 0
            }
            
            self.parking_records.append(record)
            self._save_to_csv(record)
            self.update_table()
            
            msg = f"✅ CHECK IN SUCCESSFUL\n\nCard ID: {card_id}\nLicense Plate: {self.captured_license_plate}\nSlot: {slot}"
            print(f"\n✅ AUTO CHECK-IN: {card_id} | Plate: {self.captured_license_plate} | Slot: {slot} | Time: {now.strftime('%H:%M:%S')}")
            self._update_status(f"Status: ✓ {card_id} checked in successfully")
            
            # Clear temporary data
            self.pending_card_data = None
            self.is_processing_card = False
            
        except Exception as e:
            print(f"❌ Auto check-in error: {e}")
            self._update_status(f"Status: Auto check-in error - {str(e)}")
            self.is_processing_card = False
    
    def _auto_check_out(self):
        """Auto check-out after successful license plate extraction"""
        try:
            if not self.pending_card_data or not self.captured_license_plate:
                self.is_processing_card = False
                return
            
            license_plate = self.captured_license_plate.upper()
            card_id = self.pending_card_data['card_id']
            
            # Find existing record
            record = None
            for r in self.parking_records:
                if r['license_plate'] == license_plate:
                    record = r
                    break
            
            if not record:
                self._update_status(f"Status: No vehicle with plate {license_plate} found")
                print(f"❌ License plate {license_plate} not found in records")
                self.is_processing_card = False
                return
            
            # Check if already checked out
            if record['status'] == 'OUT':
                self._update_status(f"Status: Vehicle {license_plate} already checked out")
                self.is_processing_card = False
                return
            
            # Calculate parking fee
            time_in = record['time_in']
            time_out = datetime.now()
            duration = time_out - time_in
            hours = duration.total_seconds() / 3600
            
            # Fee calculation: < 2h = 20000 VND, +10000 per extra hour
            if hours <= 2:
                fee = 20000
            else:
                fee = 20000 + int((hours - 2) * 10000)
            
            # Update record with checkout info
            record['time_out'] = time_out
            record['status'] = 'OUT'
            record['fee'] = fee
            
            self._save_to_csv(record)
            self.update_table()
            
            if self.fee_label:
                self.fee_label.setText(f"💰 PARKING FEE\n{fee:,} VND")
            
            print(f"\n✅ AUTO CHECK-OUT: {license_plate} | Card: {card_id} | Slot: {record['slot']} | Fee: {fee:,} VND | Duration: {hours:.2f}h")
            self._update_status(f"Status: ✓ {license_plate} checked out | Fee: {fee:,} VND")
            
            # Clear temporary data
            self.pending_card_data = None
            self.is_processing_card = False
            
        except Exception as e:
            print(f"❌ Auto check-out error: {e}")
            self._update_status(f"Status: Auto check-out error - {str(e)}")
            self.is_processing_card = False
    
    def on_frame_received(self, frame):
        """Display incoming camera frame in UI"""
        self.current_frame = frame
        try:
            rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            h, w, ch = rgb_frame.shape
            bytes_per_line = 3 * w
            qt_image = QImage(rgb_frame.data, w, h, bytes_per_line, QImage.Format_RGB888)
            pixmap = QPixmap.fromImage(qt_image)
            scaled = pixmap.scaledToWidth(700, Qt.SmoothTransformation)
            self.camera_label.setPixmap(scaled)
        except Exception as e:
            print(f"Frame display error: {e}")
    
    def on_camera_error(self, msg):
        """Handle and display camera errors to user"""
        print(f"❌ Camera Error: {msg}")
        self._update_status(f"Status: Camera Error - {msg}")
    
    def capture_plate(self):
        """Manually capture current frame and extract license plate via YOLO + OCR"""
        if self.current_frame is None:
            QMessageBox.warning(self, "Error", "No frame to capture")
            return
        
        try:
            if not self.camera_thread.license_plate_detector:
                QMessageBox.warning(self, "Error", "YOLO model not loaded")
                return
            
            self._update_status("Status: Processing...")
            
            # YOLO detection
            detections = self.camera_thread.license_plate_detector(self.current_frame)[0]
            
            if len(detections.boxes) == 0:
                QMessageBox.information(self, "Result", "No license plate detected")
                self._update_status("Status: No plate detected")
                return
            
            # Get first detection
            detection = detections.boxes.data.tolist()[0]
            x1, y1, x2, y2, score, class_id = detection
            
            # Crop license plate area
            crop = self.current_frame[int(y1):int(y2), int(x1):int(x2), :]
            
            # Display snapshot (cropped license plate image)
            rgb_crop = cv2.cvtColor(crop, cv2.COLOR_BGR2RGB)
            h, w, ch = rgb_crop.shape
            bytes_per_line = 3 * w
            qt_image = QImage(rgb_crop.data, w, h, bytes_per_line, QImage.Format_RGB888)
            pixmap = QPixmap.fromImage(qt_image)
            scaled_pixmap = pixmap.scaledToWidth(380, Qt.SmoothTransformation)
            self.snapshot_label.setPixmap(scaled_pixmap)
            
            # Save FULL FRAME (entire video capture) to captured_images folder
            try:
                # Create captured_images directory if it doesn't exist
                captured_dir = Path(__file__).parent / "captured_images"
                captured_dir.mkdir(exist_ok=True)
                
                # Generate unique filename with timestamp
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")[:-3]  # Remove last 3 digits for milliseconds
                
                # Save FULL FRAME (entire screenshot with entire vehicle/parking area)
                full_frame_filename = f"full_frame_{timestamp}.jpg"
                full_frame_path = captured_dir / full_frame_filename
                cv2.imwrite(str(full_frame_path), self.current_frame)
                print(f"✓ Full frame saved: {full_frame_path}")
                
                # Also save cropped license plate (for reference/verification)
                crop_filename = f"license_plate_{timestamp}.jpg"
                crop_path = captured_dir / crop_filename
                cv2.imwrite(str(crop_path), crop)
                print(f"✓ License plate crop saved: {crop_path}")
            except Exception as save_error:
                print(f"Warning: Could not save images: {save_error}")
            
            # OCR extraction
            plate_text, conf = extract_license_plate(crop)
            
            if not plate_text:
                QMessageBox.information(self, "Result", "Could not extract license plate text")
                self._update_status("Status: OCR failed - no text recognized")
                self.extracted_plate_label.setText("OCR Failed")
                return
            
            # Store the extracted license plate
            self.captured_license_plate = plate_text
            
            # Display extracted license plate (purple box)
            self.extracted_plate_label.setText(plate_text)
            self.extracted_plate_label.setStyleSheet(
                "background-color: #f3e8ff; color: #6d28d9; padding: 10px; "
                "border-radius: 6px; border: 2px solid #7c3aed; letter-spacing: 2px; "
                "font-size: 16pt; font-weight: bold;"
            )
            
            # Update status
            self._update_status(f"Status: ✓ License plate captured - {plate_text}")
            
            # Update Time In display
            current_time = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
            self.time_in_display.setText(current_time)
            
            print(f"\n✅ License Plate Captured: {plate_text} (Confidence: {conf:.2%})")
            print(f"   Time In: {current_time}")
        
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Capture error: {str(e)}")
            print(f"Capture error: {e}")
            import traceback
            traceback.print_exc()
    
    def check_in(self):
        """Manual check-in: Record vehicle entry with card ID, slot, and captured license plate"""
        try:
            card_id = self.card_id_input.text().strip()
            slot = self.slot_input.text().strip()
            
            if not card_id or not slot:
                QMessageBox.warning(self, "Error", "Please enter Card ID and Slot Number")
                return
            
            if not self.captured_license_plate:
                QMessageBox.warning(self, "Error", "Please capture license plate first")
                return
            
            # Check if vehicle already checked in
            for record in self.parking_records:
                if record['card_id'] == card_id and record['status'] == 'IN':
                    QMessageBox.warning(self, "Error", f"Vehicle already checked in at slot {record['slot']}")
                    return
            
            now = datetime.now()
            
            # Show confirmation message
            msg_text = (
                f"<b>✓ CONFIRM CHECK IN</b><br><br>"
                f"<b>Card ID:</b> {card_id}<br>"
                f"<b>License Plate:</b> {self.captured_license_plate}<br>"
                f"<b>Parking Slot:</b> {slot}<br>"
                f"<b>Time:</b> {now.strftime('%H:%M:%S')}<br><br>"
                f"<b>Do you want to proceed?</b>"
            )
            
            # Show custom confirmation dialog
            dialog = CheckInConfirmationDialog(self, card_id, self.captured_license_plate, slot, now.strftime("%d/%m/%Y %H:%M:%S"))
            
            if dialog.exec_() != QDialog.Accepted:
                self._update_status("Status: Check-in cancelled by user")
                print(f"⚠️  Check-in cancelled for card {card_id}")
                return
            
            # Create new parking record with captured license plate
            record = {
                'card_id': card_id,
                'license_plate': self.captured_license_plate,  # Use extracted license plate from OCR
                'time_in': now,
                'time_out': None,
                'slot': slot,
                'status': 'IN',
                'fee': 0
            }
            
            self.parking_records.append(record)
            
            # Save record to CSV
            try:
                self._save_to_csv(record)
            except Exception as csv_error:
                print(f"⚠️  Warning: CSV save failed: {csv_error}")
                # Continue even if CSV save fails
            
            # Refresh table display
            try:
                self.update_table()
            except Exception as table_error:
                print(f"⚠️  Warning: Table update failed: {table_error}")
                # Continue even if table update fails
            
            # Show success message
            msg = f"✅ Check-in Successful\n\nCard ID: {card_id}\nLicense Plate: {self.captured_license_plate}\nSlot: {slot}\nTime: {now.strftime('%d/%m/%Y %H:%M:%S')}"
            QMessageBox.information(self, "Check-in Success", msg)
            
            self._update_status(f"Status: Checked in - Card {card_id}, Plate {self.captured_license_plate}")
            
            # Clear input fields for next transaction
            self.card_id_input.clear()
            self.slot_input.clear()
            
            print(f"\n✅ CHECK IN - Card: {card_id}, Slot: {slot}, Plate: {self.captured_license_plate}, Time: {now.strftime('%Y-%m-%d %H:%M:%S')}")
        
        except Exception as e:
            print(f"❌ Check-in error: {e}")
            import traceback
            traceback.print_exc()
            QMessageBox.critical(self, "Error", f"Check-in failed: {str(e)}")
    
    def check_out(self):
        """Manual check-out: Record vehicle exit, calculate fee, and update status"""
        try:
            license_plate = self.checkout_plate_input.text().strip().upper()
            
            if not license_plate:
                QMessageBox.warning(self, "Error", "Please enter License Plate")
                return
            
            # Search for vehicle record by license plate
            record = None
            for r in self.parking_records:
                if r['license_plate'] == license_plate:
                    record = r
                    break
            
            if not record:
                msg = "❌ No vehicle with the above license plate was found."
                QMessageBox.warning(self, "Not Found", msg)
                print(f"❌ License plate not found: {license_plate}")
                return
            
            # Check if already checked out
            if record['status'] == 'OUT':
                QMessageBox.warning(self, "Error", "Vehicle already checked out")
                return
            
            # Calculate parking duration and fee
            time_in = record['time_in']
            time_out = datetime.now()
            duration = time_out - time_in
            hours = duration.total_seconds() / 3600
            
            # Fee structure: < 2 hours = 20000 VND, +10000 VND per additional hour
            if hours <= 2:
                fee = 20000
            else:
                fee = 20000 + int((hours - 2) * 10000)
            
            # Show confirmation dialog with custom dialog
            dialog = CheckOutConfirmationDialog(self, license_plate, record['slot'], hours, fee, time_out.strftime("%d/%m/%Y %H:%M:%S"))
            
            if dialog.exec_() != QDialog.Accepted:
                self._update_status("Status: Check-out cancelled by user")
                print(f"⚠️  Check-out cancelled for plate {license_plate}")
                return
            
            # Update record with checkout information
            record['time_out'] = time_out
            record['status'] = 'OUT'
            record['fee'] = fee
            
            # Update Time Out display
            checkout_time = time_out.strftime("%d/%m/%Y %H:%M:%S")
            self.time_out_display.setText(checkout_time)
            
            # Save to CSV file
            try:
                self._save_to_csv(record)
            except Exception as csv_error:
                print(f"⚠️  Warning: CSV save failed: {csv_error}")
                # Continue even if CSV save fails
            
            # Refresh table display
            try:
                self.update_table()
            except Exception as table_error:
                print(f"⚠️  Warning: Table update failed: {table_error}")
                # Continue even if table update fails
            
            # Show success message
            msg = f"✅ Check-out Successful\n\nLicense Plate: {license_plate}\nSlot: {record['slot']}\nFee: {fee:,} VND\nDuration: {hours:.2f} hours\nTime: {time_out.strftime('%d/%m/%Y %H:%M:%S')}"
            QMessageBox.information(self, "Check-out Success", msg)
            
            # Update status and fee display
            self._update_status(f"Status: Vehicle {license_plate} checked out | Parking slot: {record['slot']}")
            try:
                self.fee_label.setText(f"💰 PARKING FEE\n{fee:,} VND")
            except Exception as fee_error:
                print(f"⚠️  Warning: Fee label update failed: {fee_error}")
            
            self.checkout_plate_input.clear()
            
            print(f"\n✅ CHECK OUT - Plate: {license_plate}, Slot: {record['slot']}, Fee: {fee:,} VND")
        
        except Exception as e:
            print(f"❌ Check-out error: {e}")
            import traceback
            traceback.print_exc()
            QMessageBox.critical(self, "Error", f"Check-out failed: {str(e)}")
    
    def update_slot_status(self):
        """Update parking slot display based on current records"""
        # Clear all slots (set to empty/green)
        for slot_info in self.slot_buttons:
            slot_num = slot_info['slot_num']
            slot_info['is_occupied'] = False
            slot_info['button'].setStyleSheet(
                "QPushButton {"
                "  background-color: #10b981;"
                "  color: white;"
                "  font-weight: bold;"
                "  border-radius: 6px;"
                "  border: 2px solid #059669;"
                "  padding: 6px;"
                "}"
                "QPushButton:hover {"
                "  background-color: #059669;"
                "}"
            )
        
        # Mark occupied slots (red) for vehicles currently parked
        for record in self.parking_records:
            if record['status'] == 'IN':  # Vehicle is currently parked
                slot_num = int(record['slot'])
                if 1 <= slot_num <= 10:
                    for slot_info in self.slot_buttons:
                        if slot_info['slot_num'] == slot_num:
                            slot_info['is_occupied'] = True
                            slot_info['button'].setStyleSheet(
                                "QPushButton {"
                                "  background-color: #ef4444;"
                                "  color: white;"
                                "  font-weight: bold;"
                                "  border-radius: 6px;"
                                "  border: 2px solid #dc2626;"
                                "  padding: 6px;"
                                "}"
                                "QPushButton:hover {"
                                "  background-color: #dc2626;"
                                "}"
                            )
                            break
    
    def update_table(self):
        """Refresh parking records table with current data"""
        try:
            self.table.setRowCount(0)
            
            for i, record in enumerate(self.parking_records):
                self.table.insertRow(i)
                
                # Populate table cells
                self.table.setItem(i, 0, QTableWidgetItem(record['card_id']))
                self.table.setItem(i, 1, QTableWidgetItem(record['license_plate']))
                self.table.setItem(i, 2, QTableWidgetItem(record['time_in'].strftime("%d/%m/%Y %H:%M:%S") if record['time_in'] else ""))
                self.table.setItem(i, 3, QTableWidgetItem(record['time_out'].strftime("%d/%m/%Y %H:%M:%S") if record['time_out'] else ""))
                self.table.setItem(i, 4, QTableWidgetItem(record['slot']))
                
                # Color-code status column
                status_item = QTableWidgetItem(record['status'])
                if record['status'] == 'IN':
                    status_item.setBackground(QColor("#c8e6c9"))  # Green for entry
                else:
                    status_item.setBackground(QColor("#ffccbc"))  # Orange for exit
                self.table.setItem(i, 5, status_item)
                
                # Display fee if checkout occurred
                self.table.setItem(i, 6, QTableWidgetItem(str(record['fee']) if record['fee'] > 0 else ""))
            
            # Update parking slots display based on current records
            self.update_slot_status()
        except Exception as e:
            print(f"❌ Error updating table: {e}")
            import traceback
            traceback.print_exc()
    
    def _save_to_csv(self, record):
        """Append parking record to CSV file"""
        try:
            with open(self.csv_file, 'a', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow([
                    record['card_id'],
                    record['license_plate'],
                    record['time_in'].strftime("%Y-%m-%d %H:%M:%S") if record['time_in'] else "",
                    record['time_out'].strftime("%Y-%m-%d %H:%M:%S") if record['time_out'] else "",
                    record['slot'],
                    record['status'],
                    record['fee'] if record['fee'] > 0 else ""
                ])
        except Exception as e:
            print(f"❌ Error saving to CSV: {e}")
    
    def export_to_excel(self):
        """Export all parking records to Excel file with formatting"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            # Get file path from dialog
            file_path, _ = QFileDialog.getSaveFileName(
                self, 
                "Export Parking Records to Excel", 
                f"parking_records_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                "Excel Files (*.xlsx);;All Files (*)"
            )
            
            if not file_path:
                return
            
            # Ensure .xlsx extension
            if not file_path.endswith('.xlsx'):
                file_path = file_path + '.xlsx'
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Parking Records"
            
            # Add metadata
            ws['A1'].value = "Parking Management System - Records Export"
            ws['A2'].value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws['A2'].font = Font(italic=True, size=9)
            ws['A3'].value = f"Total Records: {len(self.parking_records)}"
            ws['A3'].font = Font(bold=True, size=10)
            
            # Headers (starting from row 5)
            headers = ['Card ID', 'License Plate', 'Time In', 'Time Out', 'Slot', 'Status', 'Fee (VND)']
            ws.append([])  # Row 4 - blank
            ws.append(headers)  # Row 5
            
            # Header style
            header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for cell in ws[5]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
            
            # Add data rows
            for record in self.parking_records:
                time_in_str = record['time_in'].strftime("%Y-%m-%d %H:%M:%S") if record['time_in'] else ""
                time_out_str = record['time_out'].strftime("%Y-%m-%d %H:%M:%S") if record['time_out'] else ""
                fee_str = f"{record['fee']:,}" if record['fee'] > 0 else ""
                
                row = [
                    record['card_id'],
                    record['license_plate'],
                    time_in_str,
                    time_out_str,
                    record['slot'],
                    record['status'],
                    fee_str
                ]
                ws.append(row)
            
            # Format data cells
            for row in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=1, max_col=7):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    
                    # Format fee column (right-align for numbers)
                    if cell.column == 7:  # Fee column
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    
                    # Color status cells
                    if cell.column == 6:  # Status column
                        if cell.value == 'IN':
                            cell.fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
                            cell.font = Font(bold=True, color="155724")
                        elif cell.value == 'OUT':
                            cell.fill = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")
                            cell.font = Font(bold=True, color="721c24")
            
            # Set column widths for better readability
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 22
            ws.column_dimensions['D'].width = 22
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 18
            
            # Freeze header row for easier scrolling
            ws.freeze_panes = 'A6'
            
            # Save the workbook to file
            wb.save(file_path)
            
            # Calculate summary statistics
            total_in = sum(1 for r in self.parking_records if r['status'] == 'IN')
            total_out = sum(1 for r in self.parking_records if r['status'] == 'OUT')
            total_fee = sum(r['fee'] for r in self.parking_records if r['fee'] > 0)
            
            # Show success message
            success_msg = (
                f"✅ Excel file saved successfully!\n\n"
                f"File: {file_path}\n"
                f"Total Records: {len(self.parking_records)}\n"
                f"Vehicles IN: {total_in}\n"
                f"Vehicles OUT: {total_out}\n"
                f"Total Revenue: {total_fee:,} VND"
            )
            QMessageBox.information(self, "Export Success", success_msg)
            print(f"✅ Exported to Excel: {file_path}")
            print(f"   Records: {len(self.parking_records)} | IN: {total_in} | OUT: {total_out} | Revenue: {total_fee:,} VND")
        
        except ImportError:
            QMessageBox.critical(
                self, 
                "Error", 
                "openpyxl module not installed.\n\n"
                "Please install it using:\n"
                "pip install openpyxl\n\n"
                "Or check requirements.txt and run:\n"
                "pip install -r requirements.txt"
            )
            print("❌ openpyxl not installed")
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed to export Excel:\n\n{str(e)}")
            print(f"❌ Export error: {e}")
            import traceback
            traceback.print_exc()
    
    def on_home_button_clicked(self):
        """Handle HOME button click - scroll back to top of right panel"""
        # Find the scroll area in the right panel and scroll to top
        try:
            # Scroll to top of the scroll area
            scroll_areas = self.findChildren(QScrollArea)
            if scroll_areas:
                scroll_area = scroll_areas[0]  # Get the first scroll area (right panel)
                scroll_bar = scroll_area.verticalScrollBar()
                scroll_bar.setValue(0)  # Scroll to top
                
                # Show confirmation in status
                self._update_status("Status: Scrolled to top")
                print("🏠 HOME button clicked - scrolled to top")
        except Exception as e:
            print(f"⚠️  Error on HOME button click: {e}")
    
    def closeEvent(self, event):
        """Clean up resources when application exits"""
        # Stop serial port thread
        if self.serial_thread:
            self.serial_thread.stop()
            self.serial_thread.join(timeout=2)
        
        # Stop camera thread
        if self.camera_thread:
            self.camera_thread.stop()
        
        event.accept()


def main():
    """Main application entry point - Initialize PyQt5 application and show UI"""
    print("=" * 70)
    print("🅿️  PARKING MANAGEMENT SYSTEM")
    print("=" * 70)
    print()
    
    app = QApplication(sys.argv)
    app.setStyle('Fusion')
    
    try:
        # Create and display main window
        window = ParkingManagementUI()
        window.show()
        
        print("✓ Window displayed")
        print("✓ Application ready\n")
        
        # Start Qt event loop
        sys.exit(app.exec_())
    
    except Exception as e:
        print(f"✗ Error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
